#!/usr/bin/env python3
"""
估值 Excel 模型生成脚本
IR step6b_valuation 和 BP bp_估值 共用

用法（IR）:
  python3 build_valuation_excel.py --pipeline ir --task-id TASK-XXXXX

用法（BP）:
  python3 build_valuation_excel.py --pipeline bp --task-id TASK-XXXXX

输出:
  {IR_RUNTIME}/jobs/{JOB_ID}/outputs/{TASK_ID}_valuation.xlsx
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── 配色规范 (aRGB: 8-digit hex with alpha prefix) ──
DARK_BLUE = "FF1F4E79"
LIGHT_BLUE = "FFD9E1F2"
MED_BLUE = "FFBDD7EE"
LIGHT_GRAY = "FFF2F2F2"
WHITE_FONT = "FFFFFFFF"
BLUE_FONT = "FF0000FF"
BLACK_FONT = "FF000000"
GREEN_FONT = "FF008000"

FILL_DARK_BLUE = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_MED_BLUE = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

FONT_HEADER = Font(name="Calibri", size=12, bold=True, color=WHITE_FONT)
FONT_COL_HEADER = Font(name="Calibri", size=11, bold=True, color=BLACK_FONT)
FONT_INPUT = Font(name="Calibri", size=11, color=BLUE_FONT)
FONT_FORMULA = Font(name="Calibri", size=11, color=BLACK_FONT)
FONT_LINK = Font(name="Calibri", size=11, color=GREEN_FONT)
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=BLACK_FONT)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=WHITE_FONT)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="FFB4B2A9")
)

IR_RUNTIME = os.path.expanduser("~/.workbuddy/ir_runtime")


def apply_section_header(ws, row, cols, title):
    """Apply dark blue section header across columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_TITLE
    cell.fill = FILL_DARK_BLUE
    cell.alignment = ALIGN_LEFT


def apply_col_headers(ws, row, headers):
    """Apply light blue column headers."""
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_LIGHT_BLUE
        cell.alignment = ALIGN_CENTER


def set_col_widths(ws, widths):
    """Set column widths."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_ir_dcf_sheet(ws, data):
    """Build DCF sheet for IR pipeline."""
    apply_section_header(ws, 1, 9, "DCF Valuation Model")
    ws.row_dimensions[1].height = 30

    headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Terminal"]
    apply_col_headers(ws, 3, headers)

    rows_data = data.get("dcf", {})
    row = 4

    # Revenue
    ws.cell(row=row, column=1, value="Revenue").font = FONT_BOLD
    rev_values = rows_data.get("revenue", [0]*5)
    for i, v in enumerate(rev_values[:5]):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.font = FONT_FORMULA
        cell.number_format = '#,##0'
        cell.alignment = ALIGN_RIGHT
    row += 1

    # Revenue Growth
    ws.cell(row=row, column=1, value="Revenue Growth %").font = FONT_BOLD
    growth = rows_data.get("revenue_growth", [0]*5)
    for i, v in enumerate(growth[:5]):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.font = FONT_FORMULA
        cell.number_format = '0.0%'
        cell.alignment = ALIGN_RIGHT
    row += 1

    # EBITDA
    ws.cell(row=row, column=1, value="EBITDA").font = FONT_BOLD
    ebitda = rows_data.get("ebitda", [0]*5)
    for i, v in enumerate(ebitda[:5]):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.font = FONT_FORMULA
        cell.number_format = '#,##0'
        cell.alignment = ALIGN_RIGHT
    row += 1

    # EBITDA Margin
    ws.cell(row=row, column=1, value="EBITDA Margin %").font = FONT_BOLD
    margins = rows_data.get("ebitda_margin", [0]*5)
    for i, v in enumerate(margins[:5]):
        cell = ws.cell(row=row, column=i+2, value=v)
        cell.font = FONT_FORMULA
        cell.number_format = '0.0%'
        cell.alignment = ALIGN_RIGHT
    row += 2

    # UFCF section
    ws.cell(row=row, column=1, value="UFCF Calculation").font = FONT_BOLD
    ws.cell(row=row, column=1).fill = FILL_MED_BLUE
    row += 1

    ufcf_items = ["EBIT × (1-Tax)", "+ D&A", "- CapEx", "- ΔWC", "= UFCF"]
    ufcf_values = rows_data.get("ufcf_detail", {})
    for item in ufcf_items:
        ws.cell(row=row, column=1, value=item).font = FONT_FORMULA
        vals = ufcf_values.get(item.replace(" ", "").replace("=", ""), [0]*5)
        for i, v in enumerate(vals[:5]):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.font = FONT_FORMULA
            cell.number_format = '#,##0'
            cell.alignment = ALIGN_RIGHT
            if item.startswith("="):
                cell.font = FONT_BOLD
                cell.border = Border(top=Side(style="thin"), bottom=Side(style="double"))
        row += 1

    row += 1
    # Discount section
    ws.cell(row=row, column=1, value="Discount to Present Value").font = FONT_BOLD
    ws.cell(row=row, column=1).fill = FILL_MED_BLUE
    row += 1

    disc_items = ["Discount Factor", "PV of UFCF"]
    for item in disc_items:
        ws.cell(row=row, column=1, value=item).font = FONT_FORMULA
        vals = rows_data.get(item.lower().replace(" ", "_"), [0]*5)
        for i, v in enumerate(vals[:5]):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.font = FONT_FORMULA
            if "factor" in item.lower():
                cell.number_format = '0.000'
            else:
                cell.number_format = '#,##0'
            cell.alignment = ALIGN_RIGHT
        row += 1

    row += 1
    # Valuation bridge
    ws.cell(row=row, column=1, value="Valuation Bridge").font = FONT_BOLD
    ws.cell(row=row, column=1).fill = FILL_MED_BLUE
    row += 1

    bridge = data.get("valuation_bridge", {})
    bridge_items = [
        ("Sum of PV of UFCF", bridge.get("pv_sum", 0)),
        ("Terminal Value", bridge.get("terminal_value", 0)),
        ("PV of Terminal Value", bridge.get("pv_terminal", 0)),
        ("Enterprise Value", bridge.get("ev", 0)),
        ("- Net Debt", bridge.get("net_debt", 0)),
        ("- Minority Interest", bridge.get("minority", 0)),
        ("+ Associates", bridge.get("associates", 0)),
        ("Equity Value", bridge.get("equity_value", 0)),
        ("Shares Outstanding", bridge.get("shares", 0)),
        ("Implied Price per Share", bridge.get("price_per_share", 0)),
    ]
    for label, val in bridge_items:
        ws.cell(row=row, column=1, value=label).font = FONT_FORMULA
        cell = ws.cell(row=row, column=2, value=val)
        cell.alignment = ALIGN_RIGHT
        if "Price" in label or "Shares" in label:
            cell.font = FONT_BOLD
            cell.fill = FILL_MED_BLUE
        else:
            cell.font = FONT_LINK
        if "Price" in label:
            cell.number_format = '#,##0.00'
        elif "Shares" in label:
            cell.number_format = '#,##0'
        else:
            cell.number_format = '#,##0'
        row += 1

    set_col_widths(ws, [25, 15, 15, 15, 15, 15, 15])


def build_ir_comps_sheet(ws, data):
    """Build Comps sheet for IR pipeline."""
    apply_section_header(ws, 1, 10, "Comparable Company Analysis")
    ws.row_dimensions[1].height = 30

    comps = data.get("comps", {})
    companies = comps.get("companies", [])
    metrics = comps.get("metrics", [])

    # Operating metrics section
    apply_section_header(ws, 3, 10, "Operating Statistics & Financial Metrics")

    op_headers = ["Company"] + [m.get("name", "") for m in metrics if m.get("type") == "operating"]
    apply_col_headers(ws, 4, op_headers)

    for i, co in enumerate(companies):
        row = 5 + i
        ws.cell(row=row, column=1, value=co.get("name", "")).font = FONT_BOLD
        for j, m in enumerate(metrics):
            if m.get("type") != "operating":
                continue
            val = co.get("values", {}).get(m["name"])
            cell = ws.cell(row=row, column=j+2, value=val)
            cell.font = FONT_INPUT if m.get("is_input") else FONT_FORMULA
            cell.alignment = ALIGN_RIGHT
            if "%" in m.get("format", ""):
                cell.number_format = '0.0%'
            elif "x" in m.get("format", ""):
                cell.number_format = '0.0"x"'
            else:
                cell.number_format = '#,##0'

    # Statistics row
    stat_row = 5 + len(companies) + 1
    stats = ["Maximum", "75th Pctl", "Median", "25th Pctl", "Minimum"]
    for s in stats:
        ws.cell(row=stat_row, column=1, value=s).font = FONT_BOLD
        ws.cell(row=stat_row, column=1).fill = FILL_LIGHT_GRAY
        # Values would be filled with actual data
        stat_row += 1

    # Valuation section
    val_row = stat_row + 2
    apply_section_header(ws, val_row, 10, "Valuation Multiples")
    val_row += 1

    val_headers = ["Company"] + [m.get("name", "") for m in metrics if m.get("type") == "valuation"]
    apply_col_headers(ws, val_row, val_headers)
    val_row += 1

    for i, co in enumerate(companies):
        row = val_row + i
        ws.cell(row=row, column=1, value=co.get("name", "")).font = FONT_BOLD
        for j, m in enumerate(metrics):
            if m.get("type") != "valuation":
                continue
            val = co.get("values", {}).get(m["name"])
            cell = ws.cell(row=row, column=j+2, value=val)
            cell.font = FONT_INPUT if m.get("is_input") else FONT_FORMULA
            cell.alignment = ALIGN_RIGHT
            if "x" in m.get("format", ""):
                cell.number_format = '0.0"x"'
            elif "%" in m.get("format", ""):
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0'

    set_col_widths(ws, [18] + [14]*9)


def build_sensitivity_sheet(ws, data):
    """Build 5x5 sensitivity matrix."""
    apply_section_header(ws, 1, 8, "Sensitivity Analysis — Implied Price per Share")
    ws.row_dimensions[1].height = 30

    sens = data.get("sensitivity", {})
    row_var = sens.get("row_variable", "Terminal Growth Rate")
    col_var = sens.get("col_variable", "WACC")
    row_vals = sens.get("row_values", [])
    col_vals = sens.get("col_values", [])
    matrix = sens.get("matrix", [])

    # Column variable header
    ws.cell(row=3, column=1, value=f"{row_var} \\ {col_var}").font = FONT_COL_HEADER
    ws.cell(row=3, column=1).fill = FILL_LIGHT_BLUE

    for j, cv in enumerate(col_vals):
        cell = ws.cell(row=3, column=j+2, value=cv)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_LIGHT_BLUE
        cell.alignment = ALIGN_CENTER
        cell.number_format = '0.0%'

    # Rows
    center_row = len(row_vals) // 2
    center_col = len(col_vals) // 2

    for i, rv in enumerate(row_vals):
        row = 4 + i
        cell = ws.cell(row=row, column=1, value=rv)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_LIGHT_BLUE
        cell.alignment = ALIGN_CENTER
        cell.number_format = '0.0%'

        for j in range(len(col_vals)):
            val = matrix[i][j] if i < len(matrix) and j < len(matrix[i]) else 0
            cell = ws.cell(row=row, column=j+2, value=val)
            cell.font = FONT_FORMULA
            cell.alignment = ALIGN_RIGHT
            cell.number_format = '#,##0.00'
            # Highlight center cell (base case)
            if i == center_row and j == center_col:
                cell.fill = FILL_MED_BLUE
                cell.font = FONT_BOLD

    set_col_widths(ws, [22] + [14]*7)


def build_bp_returns_sheet(ws, data):
    """Build Returns sheet for BP pipeline."""
    apply_section_header(ws, 1, 10, "Investment Returns Model")
    ws.row_dimensions[1].height = 30

    returns = data.get("returns", {})

    # Key assumptions
    apply_section_header(ws, 3, 10, "Key Assumptions")

    assumptions = returns.get("assumptions", {})
    row = 4
    for key, val in assumptions.items():
        ws.cell(row=row, column=1, value=key).font = FONT_BOLD
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = FONT_INPUT
        cell.alignment = ALIGN_RIGHT
        if isinstance(val, float) and val < 1:
            cell.number_format = '0.0%'
        elif isinstance(val, (int, float)):
            cell.number_format = '#,##0'
        row += 1

    row += 2
    # MOIC / IRR table
    apply_section_header(ws, row, 10, "Returns by Exit Multiple & Holding Period")
    row += 1

    exit_multiples = returns.get("exit_multiples", [3, 5, 8, 10, 15])
    hold_years = returns.get("hold_years", [3, 5, 7, 10])

    ws.cell(row=row, column=1, value="Exit Multiple \\ Years").font = FONT_COL_HEADER
    ws.cell(row=row, column=1).fill = FILL_LIGHT_BLUE

    for j, yr in enumerate(hold_years):
        cell = ws.cell(row=row, column=j+2, value=f"{yr}Y")
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_LIGHT_BLUE
        cell.alignment = ALIGN_CENTER

    row += 1
    # MOIC values
    ws.cell(row=row, column=1, value="MOIC").font = FONT_BOLD
    row += 1
    moic_matrix = returns.get("moic_matrix", [])
    for i, mult in enumerate(exit_multiples):
        ws.cell(row=row, column=1, value=f"{mult}x").font = FONT_BOLD
        for j in range(len(hold_years)):
            val = moic_matrix[i][j] if i < len(moic_matrix) and j < len(moic_matrix[i]) else 0
            cell = ws.cell(row=row, column=j+2, value=val)
            cell.font = FONT_FORMULA
            cell.number_format = '0.00"x"'
            cell.alignment = ALIGN_RIGHT
        row += 1

    row += 1
    # IRR values
    ws.cell(row=row, column=1, value="IRR").font = FONT_BOLD
    row += 1
    irr_matrix = returns.get("irr_matrix", [])
    for i, mult in enumerate(exit_multiples):
        ws.cell(row=row, column=1, value=f"{mult}x").font = FONT_BOLD
        for j in range(len(hold_years)):
            val = irr_matrix[i][j] if i < len(irr_matrix) and j < len(irr_matrix[i]) else 0
            cell = ws.cell(row=row, column=j+2, value=val)
            cell.font = FONT_FORMULA
            cell.number_format = '0.0%'
            cell.alignment = ALIGN_RIGHT
        row += 1

    set_col_widths(ws, [25] + [14]*9)


def build_assumptions_sheet(ws, data, pipeline):
    """Build Assumptions sheet."""
    label = "IR" if pipeline == "ir" else "BP"
    apply_section_header(ws, 1, 5, f"{label} Valuation — Key Assumptions")
    ws.row_dimensions[1].height = 30

    assumptions = data.get("assumptions", {})
    row = 3

    apply_col_headers(ws, row, ["Parameter", "Value", "Source", "Date"])
    row += 1

    for key, info in assumptions.items():
        if isinstance(info, dict):
            val = info.get("value", "")
            src = info.get("source", "")
            dt = info.get("date", "")
        else:
            val = info
            src = ""
            dt = ""

        ws.cell(row=row, column=1, value=key).font = FONT_BOLD
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = FONT_INPUT
        cell.alignment = ALIGN_RIGHT
        if isinstance(val, float) and val < 1:
            cell.number_format = '0.0%'
        elif isinstance(val, (int, float)):
            cell.number_format = '#,##0'

        ws.cell(row=row, column=3, value=src).font = FONT_FORMULA
        ws.cell(row=row, column=4, value=dt).font = FONT_FORMULA
        row += 1

    set_col_widths(ws, [30, 18, 35, 15])


def build_valuation_excel(pipeline, data, output_path):
    """Main builder function."""
    wb = Workbook()

    # Assumptions sheet (first)
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    build_assumptions_sheet(ws_assumptions, data, pipeline)

    if pipeline == "ir":
        # DCF sheet
        ws_dcf = wb.create_sheet("DCF")
        build_ir_dcf_sheet(ws_dcf, data)

        # Comps sheet
        ws_comps = wb.create_sheet("Comps")
        build_ir_comps_sheet(ws_comps, data)

        # Sensitivity sheet
        ws_sens = wb.create_sheet("Sensitivity")
        build_sensitivity_sheet(ws_sens, data)

    elif pipeline == "bp":
        # Returns sheet
        ws_returns = wb.create_sheet("Returns")
        build_bp_returns_sheet(ws_returns, data)

        # Comps sheet (simpler for BP)
        ws_comps = wb.create_sheet("Comps")
        build_ir_comps_sheet(ws_comps, data)

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    return output_path


def find_output_path(task_id):
    """Find the job output directory for a given task ID."""
    jobs_dir = os.path.join(IR_RUNTIME, "jobs")
    if not os.path.exists(jobs_dir):
        return None

    for job_dir in Path(jobs_dir).iterdir():
        if job_dir.is_dir():
            outputs_dir = job_dir / "outputs"
            if outputs_dir.exists():
                return str(outputs_dir / f"{task_id}_valuation.xlsx")
    return None


def main():
    parser = argparse.ArgumentParser(description="Build valuation Excel model")
    parser.add_argument("--pipeline", required=True, choices=["ir", "bp"],
                        help="Pipeline type: ir or bp")
    parser.add_argument("--task-id", required=True, help="Task ID (e.g. TASK-XXXXX)")
    parser.add_argument("--data", default=None, help="JSON data file path (optional)")
    parser.add_argument("--output", default=None, help="Output file path (optional)")
    args = parser.parse_args()

    # Load data
    if args.data and os.path.exists(args.data):
        with open(args.data, "r") as f:
            data = json.load(f)
    else:
        # Create template with empty structure
        if args.pipeline == "ir":
            data = {
                "assumptions": {
                    "Revenue Growth (Year 1)": {"value": 0.15, "source": "Management guidance", "date": ""},
                    "WACC": {"value": 0.10, "source": "CAPM calculation", "date": ""},
                    "Terminal Growth Rate": {"value": 0.03, "source": "GDP long-term", "date": ""},
                    "Tax Rate": {"value": 0.25, "source": "Statutory rate", "date": ""},
                },
                "dcf": {
                    "revenue": [0, 0, 0, 0, 0],
                    "revenue_growth": [0, 0, 0, 0, 0],
                    "ebitda": [0, 0, 0, 0, 0],
                    "ebitda_margin": [0, 0, 0, 0, 0],
                },
                "comps": {"companies": [], "metrics": []},
                "sensitivity": {
                    "row_variable": "Terminal Growth Rate",
                    "col_variable": "WACC",
                    "row_values": [0.01, 0.02, 0.03, 0.04, 0.05],
                    "col_values": [0.08, 0.09, 0.10, 0.11, 0.12],
                    "matrix": [[0]*5 for _ in range(5)],
                },
                "valuation_bridge": {},
            }
        else:
            data = {
                "assumptions": {
                    "Investment Amount": {"value": 0, "source": "", "date": ""},
                    "Entry Valuation": {"value": 0, "source": "", "date": ""},
                },
                "returns": {
                    "assumptions": {},
                    "exit_multiples": [3, 5, 8, 10, 15],
                    "hold_years": [3, 5, 7, 10],
                    "moic_matrix": [[0]*4 for _ in range(5)],
                    "irr_matrix": [[0]*4 for _ in range(5)],
                },
                "comps": {"companies": [], "metrics": []},
            }

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        found = find_output_path(args.task_id)
        if found:
            output_path = found
        else:
            output_path = os.path.join(
                IR_RUNTIME, "jobs", "default", "outputs",
                f"{args.task_id}_valuation.xlsx"
            )

    result = build_valuation_excel(args.pipeline, data, output_path)
    print(f"OK: Valuation Excel saved to {result}")


if __name__ == "__main__":
    main()
